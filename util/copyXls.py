# -*- coding:utf-8 -*-
# author:peace
# datetime:2018/9/29 17:53
# file:copyXls
# desc: 复制xls，填写测试结果工具类

import openpyxl


def copyXls(xlsPath1, xlsPath2):
    '''复制xls，复制xlsPath1保存为xlsPath2'''
    wb2 = openpyxl.Workbook()
    wb2.save(xlsPath2)
    # 打开测试数据文件、测试结果文件
    wb1 = openpyxl.load_workbook(xlsPath1, read_only=True)
    wb2 = openpyxl.load_workbook(xlsPath2)

    # 取到所有sheet的名字
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames

    # 根据第一个xls的sheet创建第二个的，创建一摸一样的
    for i in sheets1:
        wb2.create_sheet(i)
    # 删掉原来的第二个xls的sheet
    for j in sheets2:
        del wb2[j]
    # 保存第二个xls
    wb2.save(xlsPath2)

    # 遍历所有sheet，读第一个xls的数据，写入第二个中
    for i in sheets1:
        sheet1 = wb1[i]
        sheet2 = wb2[i]

        # 测试数据文件中的行数、列数
        rowCount = sheet1.max_row
        colCount = sheet1.max_column

        for m in list(range(1, rowCount + 1)):
            for n in list(range(97, 97 + colCount)):  # chr(97)='a'
                n = chr(n)  # ASCII字符
                i = '%s%d' % (n, m)  # 单元格编号
                cell1 = sheet1[i].value  # 获取data单元格数据
                sheet2[i].value = cell1  # 赋值到test单元格

    wb2.save(xlsPath2)  # 保存数据
    wb1.close()  # 关闭excel
    wb2.close()


class writeXls(object):
    '''修改excel数据'''

    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.load_workbook(self.filename)
        # self.ws = self.wb[sheetname]
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)

    def write2(self, sheetname, row_n, col_n, value):
        '''写入数据'''
        ws1 = self.wb[sheetname]
        ws1.cell(row_n, col_n).value = value
        self.wb.save(self.filename)


if __name__ == '__main__':
    # copyXls('../data/case_process.xlsx','../report/case_process_result.xlsx')

    wx = writeXls('../report/case_process_result.xlsx')
    wx.write2('Sheet1', 9, 6, 'test')

    # wt = writeXls('../report/case1_result.xlsx')
    # wt.write(3,1,'shp')

    '''测试复制sheet功能'''
    # sheets = ['Sheet1','Sheet2','Sheet3']
    # wb2 = openpyxl.Workbook()
    # wb2.save('../report/case1_result.xlsx')
    # sheets2 = wb2.sheetnames
    # print(sheets2)
    # for i in sheets:
    #     wb2.create_sheet()
    #
    # for j in sheets2:
    #     del wb2[j]
    #
    # wb2.save('../report/case1_result.xlsx')
    #
    # wb2.close()
